from openpyxl import Workbook


def create_excel():
    wb = Workbook()

    ws = wb.active
    ws.title = "Real estate data(rent)"
    headers = [
        "Title",
        "Location",
        "Rooms",
        "Shower Rooms",
        "Area",
        "Type",
        "Housing stock",
        "Price",
        "Floor",
        "Heating",
        "Destination",
        "URL",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    return wb, ws


def save_to_excel(ws, row, data):
    for col, value in enumerate(data, start=1):
        ws.cell(row=row, column=col, value=value)
